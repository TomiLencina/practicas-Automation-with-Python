import pandas as pd
from openpyxl import load_workbook

# Leer el archivo Excel original
df = pd.read_excel('202_Rpa_32569-170523407-60001.xlsx')

# Realizar las modificaciones necesarias al DataFrame
if "ID Suscription" not in df.columns:
    df.insert(loc=12, column="ID Suscription", value="")
df = df.rename(columns={"Número cuenta bancaria": "Número de cuenta bancaria"})
df.iloc[1:, [13, -2]] = ""

# Crear un archivo Excel modificado
writer = pd.ExcelWriter('ReporteNomina_modificado.xlsx', engine='openpyxl')
# Cargar el libro de trabajo del archivo Excel original
writer.book = load_workbook('202_Rpa_32569-170523407-60001.xlsx')
# Copiar el formato de las hojas del libro de trabajo original al nuevo archivo
for sheetname in writer.book.sheetnames:
    sheet = writer.book[sheetname]
    reader = pd.read_excel('202_Rpa_32569-170523407-60001.xlsx', sheet_name=sheetname)
    for idx, row in reader.iterrows():
        for cell in sheet[idx+1]:
            cell._style = row[cell.column_letter].style

# Guardar el DataFrame modificado en el nuevo archivo Excel
df.to_excel(writer, index=False, sheet_name='Sheet1')
writer.save()