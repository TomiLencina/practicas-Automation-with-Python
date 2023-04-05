import os
import time
import pandas as pd
import psutil
import pytest
#import win32com.client as win32
from openpyxl import load_workbook


class TestModificarExcel:
    @pytest.mark.modificar
    def test_modificar_excel(self):
        """ruta_carpeta = "Excels_ReporteNomina"
        lista_archivos = os.listdir(ruta_carpeta)

        for excel in lista_archivos:
            # Cerrar el archivo Excel si está abierto
             for proc in psutil.process_iter():
                archivos_abiertos = proc.open_files()
                try:
                    for archivo in archivos_abiertos:
                        if archivo.path == os.path.abspath(excel):
                            proc.kill()
                            time.sleep(2)
        # Esperar 2 segundos para asegurarse de que se haya cerrado el proceso
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    print('hoola')
                    pass
        #Abrir Excel y el archivo
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                wb = excel.Workbooks.Open(excel)
        # Establecer la propiedad Read-Only en False
                wb.ChangeFileAccess(1)
                wb.Save()"""

        ruta_carpeta = "Excels_ReporteNomina"
        lista_archivos = os.listdir(ruta_carpeta)

        for archivo in lista_archivos:
            if archivo.endswith('.xlsx'):
                archivo_excel = os.path.join(ruta_carpeta, archivo)
                workbook = load_workbook(archivo_excel, keep_vba=True)
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    worksheet.calculate_dimension()
                    worksheet.calculate_dimension(force=True)
                    worksheet.data_only = False
                workbook.save(archivo_excel)
        # Leer el archivo Excel y modificarlo
                df = pd.read_excel(os.path.join(ruta_carpeta, excel), sheet_name='ReporteNomina')
                ex = excel.gencache.EnsureDispatch('Excel.Application')
                ex.Visible = False
                if "ID Suscription" not in df.columns:
                    df.insert(loc=12, column="ID Suscription", value="")
                df = df.rename(columns={"Numero de cuenta": "Número de cuenta bancaria"})
                df.iloc[1:, [-3]] = ""
                # Guardar el DataFrame modificado en el mismo archivo Excel
                writer = pd.ExcelWriter(os.path.join(ruta_carpeta, excel), engine='openpyxl', mode='a')
                writer.book = load_workbook(os.path.join(ruta_carpeta, excel))
    # #df.(writer, index=False, sheet_name='ReporteNomina')
                writer.save()
                writer.close()